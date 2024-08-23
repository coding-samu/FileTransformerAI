from utils.utils_file import save_file_docx, save_file_txt, save_file_pdf
from ai_model.ai_model_pdf import PDFPNG, PDFJPG, PDFSVG

import os
import subprocess

from openpyxl import load_workbook

class XLSXPDF:
    def __init__(self):
        pass

    def convert(self, input_xlsx_path, output_pdf_path):
        try:
            # Verifica che il file di input sia un file XLSX valido
            if not input_xlsx_path.endswith('.xlsx'):
                raise Exception(f"Il file {input_xlsx_path} non è un file XLSX valido.")

            # Converti XLSX a PDF usando LibreOffice
            command = [
                "libreoffice",
                "--headless",  # Run without GUI
                "--convert-to", "pdf",  # Convert to PDF
                "--outdir", os.path.dirname(output_pdf_path),  # Output directory
                input_xlsx_path
            ]
            
            subprocess.run(command, check=True)

            # Rinominare il file PDF convertito se necessario
            generated_pdf_path = os.path.join(
                os.path.dirname(output_pdf_path), 
                os.path.splitext(os.path.basename(input_xlsx_path))[0] + ".pdf"
            )
            os.rename(generated_pdf_path, output_pdf_path)
            
            print(f"Conversione completata: {output_pdf_path}")
            return 0
        
        except subprocess.CalledProcessError as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1
        
        except Exception as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1

class XLSXJPG:
    def __init__(self):
        # Inizializza le classi per la conversione XLSX -> PDF e PDF -> JPG
        self.xlsx_to_pdf_converter = XLSXPDF()
        self.pdf_to_jpg_converter = PDFJPG()

    def convert(self, input_xlsx_path, output_jpg_path, page_number):
        try:
            # Definisci i nomi temporanei per i file
            temp_pdf_path = "temp_file/" + input_xlsx_path + '.pdf'
            temp_jpg_path = output_jpg_path  # Il percorso finale del JPG

            # Converti il XLSX in PDF
            if self.xlsx_to_pdf_converter.convert(input_xlsx_path, temp_pdf_path) != 0:
                raise Exception(f"Errore durante la conversione del file XLSX in PDF: {input_xlsx_path}")

            # Converti il PDF in JPG
            if self.pdf_to_jpg_converter.convert(temp_pdf_path, temp_jpg_path, page_number) != 0:
                raise Exception(f"Errore durante la conversione del file PDF in JPG: {temp_pdf_path}")

            print(f"Conversione completata: {output_jpg_path}")
            return 0

        except Exception as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1
        
        finally:
            # Rimuove il file PDF temporaneo se esiste
            if os.path.exists(temp_pdf_path):
                try:
                    os.remove(temp_pdf_path)
                    print(f"File PDF temporaneo rimosso: {temp_pdf_path}")
                except Exception as e:
                    print(f"Errore durante la rimozione del file PDF temporaneo: {e}")

class XLSXPNG:
    def __init__(self):
        # Inizializza le classi per la conversione XLSX -> PDF e PDF -> PNG
        self.xlsx_to_pdf_converter = XLSXPDF()
        self.pdf_to_png_converter = PDFPNG()

    def convert(self, input_xlsx_path, output_png_path, page_number):
        try:
            # Definisci i nomi temporanei per i file
            temp_pdf_path = "temp_file/" + input_xlsx_path + '.pdf'
            temp_png_path = output_png_path  # Il percorso finale del PNG

            # Converti il XLSX in PDF
            if self.xlsx_to_pdf_converter.convert(input_xlsx_path, temp_pdf_path) != 0:
                raise Exception(f"Errore durante la conversione del file XLSX in PDF: {input_xlsx_path}")

            # Converti il PDF in PNG
            if self.pdf_to_png_converter.convert(temp_pdf_path, temp_png_path, page_number) != 0:
                raise Exception(f"Errore durante la conversione del file PDF in PNG: {temp_pdf_path}")

            print(f"Conversione completata: {output_png_path}")
            return 0

        except Exception as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1
        
        finally:
            # Rimuove il file PDF temporaneo se esiste
            if os.path.exists(temp_pdf_path):
                try:
                    os.remove(temp_pdf_path)
                    print(f"File PDF temporaneo rimosso: {temp_pdf_path}")
                except Exception as e:
                    print(f"Errore durante la rimozione del file PDF temporaneo: {e}")

class XLSXDOCX:
    def __init__(self):
        pass

    def convert(self, input_xlsx_path, output_docx_path):
        try:
            # Carica il file XLSX
            workbook = load_workbook(input_xlsx_path)
            sheet = workbook.active

            # Prepara i dati come stringa, con ogni riga separata da una nuova linea
            docx_content = []

            # Itera su ogni riga del foglio di lavoro XLSX
            for row in sheet.iter_rows(values_only=True):
                # Unisci i valori della riga in un'unica stringa separata da spazi
                line = " ".join([str(cell) if cell is not None else "" for cell in row])
                docx_content.append(line)

            # Unisci tutte le righe in una singola stringa con newline tra di loro
            docx_data = "\n".join(docx_content)

            # Salva i dati nel file DOCX
            if save_file_docx(docx_data, output_docx_path) == 0:
                print(f"Conversione completata: {output_docx_path}")
                return 0
            else:
                raise Exception(f"Errore durante il salvataggio del file DOCX: {output_docx_path}")

        except Exception as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1

class XLSXTXT:
    pass # TODO: Implementare la conversione XLSX -> TXT

class XLSXSVG:
    def __init__(self):
        # Inizializza le classi per la conversione XLSX -> PDF e PDF -> SVG
        self.xlsx_to_pdf_converter = XLSXPDF()
        self.pdf_to_svg_converter = PDFSVG()

    def convert(self, input_xlsx_path, output_svg_path, page_number):
        try:
            # Definisci i nomi temporanei per i file
            temp_pdf_path = "temp_file/" + input_xlsx_path + '.pdf'
            temp_svg_path = output_svg_path  # Il percorso finale del SVG

            # Converti il XLSX in PDF
            if self.xlsx_to_pdf_converter.convert(input_xlsx_path, temp_pdf_path) != 0:
                raise Exception(f"Errore durante la conversione del file XLSX in PDF: {input_xlsx_path}")

            # Converti il PDF in SVG
            if self.pdf_to_svg_converter.convert_to_svg(temp_pdf_path, temp_svg_path, page_number) != 0:
                raise Exception(f"Errore durante la conversione del file PDF in SVG: {temp_pdf_path}")

            print(f"Conversione completata: {output_svg_path}")
            return 0

        except Exception as e:
            print(f"Errore durante la conversione del file {input_xlsx_path}: {e}")
            return 1
        
        finally:
            # Rimuove il file PDF temporaneo se esiste
            if os.path.exists(temp_pdf_path):
                try:
                    os.remove(temp_pdf_path)
                    print(f"File PDF temporaneo rimosso: {temp_pdf_path}")
                except Exception as e:
                    print(f"Errore durante la rimozione del file PDF temporaneo: {e}")